from tkinter import *
from tkinter import filedialog
import openpyxl as xl
import numpy as np
import datetime

file1 = None  # Global variable to store file1 workbook
file2 = None  # Global variable to store file2 workbook
file2_entry = None

# Function to upload Excel files
def upload_file(file_num):
    global file1, file2, file2_entry  # Make both file1 and file2 global variables
    file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])
    if file_path:
        if file_num == 1:
            file1_entry.delete(0, END)
            file1_entry.insert(0, file_path)
            file1 = xl.load_workbook(file_path)  # Assign the file1 workbook to the global variable
        elif file_num == 2:
            file2_entry.delete(0, END)
            file2_entry.insert(0, file_path)
            file2 = xl.load_workbook(file_path)  # Assign the file2 workbook to the global variable
            group_confirm_entry.config(state=NORMAL)
            enter_button.config(state=NORMAL)
            group_confirm_entry.config(bg="light grey")
            enter_button.config(bg="light green")

# Function to perform operations on Excel files
def perform_operations():
    global file1  # Access the global file1 workbook
    global file2  # Access the global file2 workbook
    if file2 is None or file1 is None:
        return

    # Read the specified sheets from the Excel files
    sheet1 = file1["Day on Day FC"]
    sheet2 = file2["History and Forecast Report"]

    # Add your code here to perform operations on the Excel files using openpyxl and read the sheets (sheet1 and sheet2)
    occupancy_date = []
    for row in sheet1.iter_rows(min_row=1, min_col=6, max_col=6, values_only=True):
        occupancy_date.append(row[0])

    date_array = []
    for row in sheet2.iter_rows(min_row=1, min_col=1, max_col=1, values_only=True):
        date_array.append(row[0])

    occupancy_date = np.array(occupancy_date)
    date_array = np.array(date_array)

    start = 0
    end = 0

    for i in range(len(date_array)):
        if type(date_array[i]) == datetime.datetime:
            start = i
            break

    for i in range(len(date_array)-1, -1, -1):
        if type(date_array[i]) == datetime.datetime:
            end = i
            break

    start1 = start
    start2 = start + 3

    for i in range(len(occupancy_date)):
        if type(occupancy_date[i]) == datetime.datetime:
            index = start1 if occupancy_date[i] == date_array[start1] else np.searchsorted(date_array[start2:end+1], occupancy_date[i]) + start2
            if index >= start and date_array[index] == occupancy_date[i]:
                rs_fit = int(sheet2.cell(row=index+1, column=9).value) + int(sheet2.cell(row=index+1, column=10).value)
                rs_groups = int(sheet2.cell(row=index+1, column=11).value) + int(sheet2.cell(row=index+1, column=12).value)
                rs_CH = int(sheet2.cell(row=index+1, column=6).value) + int(sheet2.cell(row=index+1, column=7).value)

                sheet1.cell(row=i+1, column=8).value = rs_fit
                sheet1.cell(row=i+1, column=9).value = rs_groups
                sheet1.cell(row=i+1, column=10).value = rs_CH

                sheet1.cell(row=i+1, column=12).value = rs_fit
                sheet1.cell(row=i+1, column=13).value = rs_groups
                sheet1.cell(row=i+1, column=14).value = rs_CH

    # Show a success message
    status_label.config(text="Operations performed successfully!")

# Function to save the modified Excel file
def save_file():
    global file1  # Access the global file1 workbook
    if file1 is None:
        return

    # Prompt the user to choose a save location and file name
    save_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Files", "*.xlsx")])

    if save_path:
        # Save the modified file with the new name and location
        file1.save(save_path)

        # Show a success message
        status_label.config(text="File saved successfully!")

# Function to update the 5th row of column K in "History and Forecast Report" sheet
def update_group_confirm():
    global file2, file2_entry
    group_confirm_value = group_confirm_entry.get()
    if file2 is not None:
        sheet2 = file2["History and Forecast Report"]
        sheet2.cell(row=5, column=11).value = group_confirm_value

        # Get the original file path
        original_path = file2_entry.get()

        # Save the modified file with the new Group Confirm value in the original path
        file2.save(original_path)

        # Show a success message
        status_label.config(text="File saved successfully!")

# Create the main window
window = Tk()
window.title("GUI Application")

# Set the window size in pixels (width x height)
window.geometry("600x400")

# Set the background color
window.configure(bg="light green")

# Add the heading label
heading_label = Label(window, text="Hotel Revenue Forecasting", font=("Lato", 12, "bold"), bg="light green")
heading_label.grid(row=0, column=0, columnspan=5, padx=10, pady=10, sticky="n")

# Create and grid a label for File 1
file1_label = Label(window, text="Forecast Report Format", font=("Lato", 10), bg="light green")
file1_label.grid(row=1, column=0, padx=10, pady=5, sticky="w")

# Create and grid an entry field for File 1
file1_entry = Entry(window, font=("Lato", 10), bg="light grey", width=25)
file1_entry.grid(row=1, column=1, padx=10, pady=5, columnspan=2)

# Create and grid a button for uploading File 1
file1_button = Button(window, text="Upload", font=("Lato", 10), bg="light green", command=lambda: upload_file(1), width=10)
file1_button.grid(row=1, column=3, padx=10, pady=5)

# Create and grid a label for File 2
file2_label = Label(window, text="History and Forecast Report", font=("Lato", 10), bg="light green")
file2_label.grid(row=2, column=0, padx=10, pady=5, sticky="w")

# Create and grid an entry field for File 2
file2_entry = Entry(window, font=("Lato", 10), bg="light grey", width=25)
file2_entry.grid(row=2, column=1, padx=10, pady=5, columnspan=2)

# Create and grid a button for uploading File 2
file2_button = Button(window, text="Upload", font=("Lato", 10), bg="light green", command=lambda: upload_file(2), width=10)
file2_button.grid(row=2, column=3, padx=10, pady=5)

# Create and grid a label for Group Confirm
group_confirm_label = Label(window, text="Group Confirm", font=("Lato", 10), bg="light green")
group_confirm_label.grid(row=3, column=0, padx=10, pady=5, sticky="w")

# Create and grid an entry field for Group Confirm
group_confirm_entry = Entry(window, font=("Lato", 10), bg="light grey", state=DISABLED, width=25)
group_confirm_entry.grid(row=3, column=1, padx=10, pady=5, columnspan=2)

# Create and grid an update button for Group Confirm
enter_button = Button(window, text="Enter", font=("Lato", 10), bg="light green", state=DISABLED, command=update_group_confirm, width=10)
enter_button.grid(row=3, column=3, padx=10, pady=5)

# Create and grid a button for performing operations
operations_button = Button(window, text="Perform Operations", font=("Lato", 10), bg="light green", command=perform_operations, width=25)
operations_button.grid(row=4, column=1, padx=10, pady=10)

# Create and grid a label for status messages
status_label = Label(window, text="", font=("Lato", 10, "bold"), bg="light green", fg="black")
status_label.grid(row=6, column=0, columnspan=4, padx=10, pady=10, sticky="w")

# Create and grid a button for saving the modified file
save_button = Button(window, text="Save File", font=("Lato", 10), bg="light green", command=save_file, width=25)
save_button.grid(row=5, column=1, padx=10, pady=10)

# Start the main window loop
window.mainloop()
